"""src/seller_console/collect_excel.py — v47 STEP5: 수집 상품 엑셀 내보내기/가져오기.

내보내기: 선택/전체 항목 → xlsx(상품ID·제목한·제목영·카테고리·가격KRW·옵션JSON·썸네일·갤러리(|)·
  상세이미지(|)·키워드·상태·수집일·원본URL). 이미지 URL은 텍스트(임베드 아님).
가져오기: 같은 템플릿 → 파싱 → 검증 리포트(신규/갱신/오류 행#·사유) → 오너 확인 후 적용(즉시 적용 아님).
  상품ID 있으면 갱신, 없으면 신규. 5000행 상한, 초과 안내. 오류는 행별(전체 롤백 아님).

openpyxl은 함수 내 지연 import(CI collect-only 안전 — Pillow 선례).
"""
from __future__ import annotations

import io
import json
from typing import Any, Dict, List, Optional, Tuple

# (키, 헤더) — 순서 = 열 순서. 내보내기/가져오기 공통 단일 소스.
COLUMNS: List[Tuple[str, str]] = [
    ("id", "상품ID"),
    ("title_ko", "제목(한)"),
    ("title_en", "제목(영)"),
    ("category", "카테고리"),
    ("price_krw", "가격(KRW)"),
    ("options", "옵션(JSON)"),
    ("thumbnail", "썸네일URL"),
    ("gallery", "갤러리URL(|구분)"),
    ("detail_images", "상세이미지URL(|구분)"),
    ("keywords", "키워드(,구분)"),
    ("status", "상태"),
    ("collected_at", "수집일"),
    ("url", "원본URL"),
]
HEADERS = [h for _, h in COLUMNS]
KEYS = [k for k, _ in COLUMNS]
MAX_ROWS = 5000

_EXAMPLE_ROW = {
    "id": "(비우면 신규 · 있으면 갱신)",
    "title_ko": "예시 상품명",
    "title_en": "Example Product",
    "category": "GEN",
    "price_krw": "12900",
    "options": '[{"name":"색상","values":["블랙","화이트"]}]',
    "thumbnail": "https://example.com/main.jpg",
    "gallery": "https://example.com/1.jpg|https://example.com/2.jpg",
    "detail_images": "https://example.com/d1.jpg",
    "keywords": "키워드1,키워드2",
    "status": "ok",
    "collected_at": "2026-07-08",
    "url": "https://example.com/product/123",
}


def _item_to_row(item: Dict[str, Any]) -> Dict[str, str]:
    """수집 이력 항목(+extra_json) → 엑셀 행 dict(전부 문자열)."""
    try:
        extra = json.loads(item.get("extra_json") or "{}")
    except Exception:
        extra = {}
    gallery = extra.get("gallery_images") or extra.get("images") or []
    if not isinstance(gallery, list):
        gallery = []
    detail = extra.get("detail_images") if isinstance(extra.get("detail_images"), list) else []
    kw = extra.get("keywords")
    if isinstance(kw, list):
        kw = ",".join(str(k) for k in kw)
    opts = extra.get("options") if isinstance(extra.get("options"), list) else []
    return {
        "id": str(item.get("id") or ""),
        "title_ko": str(extra.get("title_ko") or item.get("title") or ""),
        "title_en": str(extra.get("title_en") or extra.get("title") or ""),
        "category": str(extra.get("category_code") or extra.get("category") or ""),
        "price_krw": str(item.get("price") or extra.get("price") or ""),
        "options": json.dumps(opts, ensure_ascii=False) if opts else "",
        "thumbnail": str(item.get("image_url") or (gallery[0] if gallery else "")),
        "gallery": "|".join(str(u) for u in gallery),
        "detail_images": "|".join(str(u) for u in detail),
        "keywords": str(kw or ""),
        "status": str(item.get("status") or "ok"),
        "collected_at": str(item.get("collected_at") or "")[:19],
        "url": str(item.get("url") or ""),
    }


def build_workbook(items: List[Dict[str, Any]]) -> bytes:
    """항목 목록 → xlsx 바이트. 이미지 URL은 텍스트(임베드 아님)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "상품"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for item in items:
        row = _item_to_row(item)
        ws.append([row.get(k, "") for k in KEYS])
    # 열 너비(가독)
    widths = {"A": 14, "B": 30, "C": 30, "D": 10, "E": 12, "F": 26, "G": 32, "H": 32, "I": 24, "J": 20, "K": 8, "L": 20, "M": 34}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def template_workbook() -> bytes:
    """빈 템플릿(헤더 + 예시 1행)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "상품"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.append([_EXAMPLE_ROW.get(k, "") for k in KEYS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_workbook(file_bytes: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], bool]:
    """업로드 xlsx → (rows, errors, truncated).

    rows: 정규화된 행 dict 목록(각 dict에 _row=엑셀 행번호). errors: [{row, reason}].
    truncated: MAX_ROWS 초과로 잘렸으면 True.
    """
    from openpyxl import load_workbook

    rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    truncated = False
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return [], [{"row": 0, "reason": f"엑셀 파일을 열지 못했어요: {exc}"}], False
    ws = wb.active
    header_map: Dict[int, str] = {}
    first = True
    count = 0
    for r_idx, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if first:
            first = False
            # 헤더 → 키 매핑(순서 무관, 헤더 텍스트로 매칭)
            hdr_to_key = {h: k for k, h in COLUMNS}
            for c_idx, val in enumerate(r):
                h = str(val or "").strip()
                if h in hdr_to_key:
                    header_map[c_idx] = hdr_to_key[h]
            if "id" not in header_map.values() and "title_ko" not in header_map.values() and "url" not in header_map.values():
                return [], [{"row": 1, "reason": "헤더를 알아보지 못했어요. 템플릿을 내려받아 그 형식으로 채워 주세요."}], False
            continue
        # 빈 행 스킵
        if all((v is None or str(v).strip() == "") for v in r):
            continue
        if count >= MAX_ROWS:
            truncated = True
            break
        count += 1
        row: Dict[str, Any] = {"_row": r_idx}
        for c_idx, val in enumerate(r):
            key = header_map.get(c_idx)
            if key:
                row[key] = "" if val is None else str(val).strip()
        rows.append(row)
    return rows, errors, truncated


def _valid_price(v: str) -> Optional[str]:
    v = str(v or "").replace(",", "").strip()
    if not v:
        return ""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if f < 0:
        return None
    # 정수면 정수 문자열
    return str(int(f)) if f == int(f) else str(f)


def validate_rows(rows: List[Dict[str, Any]], existing_ids: set) -> Dict[str, Any]:
    """행 검증 → 신규/갱신/오류 분류(적용 전). 행별 오류(전체 롤백 아님).

    Returns: {new: N, update: M, errors: [{row, reason}], apply: [정규화된 적용행...]}
    apply 각 항목: {mode:'new'|'update', id, _row, fields:{...}}
    """
    new_n = 0
    update_n = 0
    errors: List[Dict[str, Any]] = []
    apply: List[Dict[str, Any]] = []
    for row in rows:
        rnum = row.get("_row", "?")
        rid = str(row.get("id") or "").strip()
        # 예시 행 안내문(비우면 신규…)은 id로 안 침
        if rid.startswith("(") or "비우면" in rid:
            rid = ""
        title_ko = str(row.get("title_ko") or "").strip()
        title_en = str(row.get("title_en") or "").strip()
        url = str(row.get("url") or "").strip()
        if not title_ko and not title_en and not url:
            errors.append({"row": rnum, "reason": "제목·원본URL이 모두 비어 있어요(빈 행)"})
            continue
        price = _valid_price(row.get("price_krw"))
        if price is None:
            errors.append({"row": rnum, "reason": f"가격이 숫자가 아니에요: {row.get('price_krw')!r}"})
            continue
        opts = row.get("options") or ""
        parsed_opts: Any = []
        if opts:
            try:
                parsed_opts = json.loads(opts)
                if not isinstance(parsed_opts, list):
                    raise ValueError
            except Exception:
                errors.append({"row": rnum, "reason": "옵션(JSON) 형식이 잘못됐어요. 예: [{\"name\":\"색상\",\"values\":[\"블랙\"]}]"})
                continue
        mode = "update" if (rid and rid in existing_ids) else "new"
        if rid and rid not in existing_ids:
            # id가 있는데 내 항목이 아님 → 오류(남의/삭제된 항목 갱신 금지)
            errors.append({"row": rnum, "reason": f"상품ID '{rid}'를 내 수집목록에서 못 찾았어요(갱신 대상 아님)"})
            continue
        if mode == "update":
            update_n += 1
        else:
            new_n += 1
        gallery = [u for u in str(row.get("gallery") or "").split("|") if u.strip()]
        detail = [u for u in str(row.get("detail_images") or "").split("|") if u.strip()]
        kw = [k for k in str(row.get("keywords") or "").split(",") if k.strip()]
        apply.append({
            "mode": mode, "id": rid, "_row": rnum,
            "fields": {
                "title_ko": title_ko, "title_en": title_en,
                "category": str(row.get("category") or "").strip(),
                "price": price,
                "options": parsed_opts,
                "thumbnail": str(row.get("thumbnail") or "").strip(),
                "gallery": gallery, "detail_images": detail,
                "keywords": kw,
                "status": str(row.get("status") or "ok").strip() or "ok",
                "url": url,
            },
        })
    return {"new": new_n, "update": update_n, "errors": errors, "apply": apply}
