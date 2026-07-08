"""tests/test_v47_excel_bulk.py — v47 STEP5: 엑셀 벌크 내보내기/가져오기.

내보내기(선택/전체) → xlsx(이미지 URL 텍스트). 가져오기 → 검증 리포트(신규/갱신/오류 행#·사유) →
확인 후 적용(즉시 적용 아님). ID 있으면 갱신·없으면 신규. 오류는 행별(전체 롤백 아님). 5000행 상한.
"""
from __future__ import annotations

import io
import json
import os
from pathlib import Path

import pytest

VIEWS = Path("src/seller_console/views.py").read_text(encoding="utf-8")
TPL = Path("src/seller_console/templates/collect_history.html").read_text(encoding="utf-8")


@pytest.fixture(autouse=True)
def _mem():
    os.environ["SELLER_CONSOLE_AUTH"] = "0"
    for k in ("DATABASE_URL", "DATABASE_URL_DIRECT"):
        os.environ.pop(k, None)
    from src.seller_console import collect_history_store as ch
    try: ch._in_memory.clear()
    except Exception: pass
    yield


# ── 모듈 ──────────────────────────────────────────────────────
def test_export_parse_roundtrip():
    from src.seller_console.collect_excel import build_workbook, parse_workbook
    items = [{"id": "abc", "title": "상품A", "price": "12900", "image_url": "https://x/t.jpg",
              "status": "ok", "collected_at": "2026-07-08T10:00:00", "url": "https://x/g-1",
              "extra_json": json.dumps({"title_ko": "상품A", "title_en": "A",
                                        "gallery_images": ["https://x/1.jpg", "https://x/2.jpg"],
                                        "options": [{"name": "색상", "values": ["블랙"]}],
                                        "category_code": "GEN", "keywords": ["kw1"]})}]
    xls = build_workbook(items)
    rows, errs, trunc = parse_workbook(xls)
    assert len(rows) == 1 and not errs and not trunc
    assert rows[0]["id"] == "abc"
    assert rows[0]["gallery"] == "https://x/1.jpg|https://x/2.jpg"     # 갤러리 | 구분, 텍스트


def test_validate_new_update_error():
    from src.seller_console.collect_excel import validate_rows
    rows = [
        {"_row": 2, "id": "mine", "title_ko": "갱신", "price_krw": "5000", "url": "u"},
        {"_row": 3, "id": "", "title_ko": "신규", "price_krw": "3000", "url": "u2"},
        {"_row": 4, "id": "", "title_ko": "나쁨", "price_krw": "abc"},        # 가격 오류
        {"_row": 5, "id": "ghost", "title_ko": "유령", "price_krw": "1"},     # 내 것 아님
    ]
    rep = validate_rows(rows, {"mine"})
    assert rep["new"] == 1 and rep["update"] == 1
    reasons = {e["row"]: e["reason"] for e in rep["errors"]}
    assert 4 in reasons and 5 in reasons


def test_template_has_header_and_example():
    from src.seller_console.collect_excel import template_workbook, parse_workbook, HEADERS
    tpl = template_workbook()
    rows, errs, _ = parse_workbook(tpl)
    assert len(rows) == 1                       # 예시 1행
    assert "제목(한)" in HEADERS and "갤러리URL(|구분)" in HEADERS


# ── 라우트 ────────────────────────────────────────────────────
def _login(c):
    with c.session_transaction() as s:
        s["user_id"] = "u1"; s["email"] = "u1"; s["authenticated"] = True


def test_export_route():
    from src.order_webhook import app
    from src.seller_console import collect_history_store as ch
    ch.append(source="extension", url="https://x/g-1", title="A", price="100", currency="KRW",
              status="ok", seller_id="u1", extra={"title_ko": "A"})
    with app.test_client() as c:
        _login(c)
        r = c.post("/seller/collect/export-xlsx", json={})
        assert r.status_code == 200
        assert "spreadsheetml" in r.mimetype
        assert len(r.data) > 0


def test_import_validate_then_apply_no_immediate():
    from src.order_webhook import app
    from src.seller_console import collect_history_store as ch
    from src.seller_console.collect_excel import build_workbook
    iid, _ = ch.append(source="extension", url="https://x/g-1", title="원본", price="100",
                       currency="KRW", status="ok", seller_id="u1", return_durable=True,
                       extra={"title_ko": "원본", "gallery_images": ["https://x/1.jpg"]})
    with app.test_client() as c:
        _login(c)
        # 내보낸 파일을 받아 제목만 바꿔 다시 올린다(갱신) + 신규 행 하나 추가
        xls = c.post("/seller/collect/export-xlsx", json={}).data
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xls)); ws = wb.active
        ws.cell(row=2, column=2, value="바뀐제목")        # 제목(한)
        ws.append(["", "새상품", "New", "GEN", "999", "", "https://x/n.jpg", "", "", "kw", "ok", "", "https://x/g-2"])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        payload = {"file": (io.BytesIO(buf.getvalue()), "p.xlsx")}
        # 1) 검증만 — 아직 적용 안 됨
        rv = c.post("/seller/collect/import-xlsx", data=payload, content_type="multipart/form-data")
        d = rv.get_json()
        assert d["ok"] and d["update"] == 1 and d["new"] == 1
        # 검증 후 원본 제목이 그대로(즉시 적용 아님)
        assert ch.get(iid, seller_ids={"u1"})["title"] == "원본"
        # 2) 적용
        ra = c.post("/seller/collect/import-apply",
                    data={"file": (io.BytesIO(buf.getvalue()), "p.xlsx")}, content_type="multipart/form-data")
        da = ra.get_json()
        assert da["ok"] and da["created"] == 1 and da["updated"] == 1
        assert ch.get(iid, seller_ids={"u1"})["title"] == "바뀐제목"    # 이제 갱신됨


def test_source_contract():
    assert "/collect/export-xlsx" in VIEWS and "/collect/import-xlsx" in VIEWS and "/collect/import-apply" in VIEWS
    assert "엑셀 내보내기" in TPL and "엑셀 업로드" in TPL
    assert "excelReportModal" in TPL and "excelApply" in TPL      # 확인 후 적용 UI
    assert "export-template" in TPL                                # 템플릿 다운로드
